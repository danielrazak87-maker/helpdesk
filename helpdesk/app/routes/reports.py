from __future__ import annotations

from flask import Blueprint, render_template, request, send_file, abort
from flask_login import login_required, current_user
from functools import wraps
from datetime import date, datetime, timedelta
from sqlalchemy import func
from typing import Any, Callable
import io
from app import db
from app.models.ticket import Ticket
from app.models.user import User
from app.models.attendance import Attendance

reports_bp = Blueprint('reports', __name__)


def admin_or_engineer(f: Callable[..., Any]) -> Callable[..., Any]:
    @wraps(f)
    def decorated(*args: Any, **kwargs: Any) -> Any:
        if not current_user.is_authenticated or current_user.is_user():
            abort(403)
        return f(*args, **kwargs)
    return decorated


@reports_bp.route('/')
@login_required
@admin_or_engineer
def index() -> str:
    return render_template('reports/index.html')


@reports_bp.route('/tickets')
@login_required
@admin_or_engineer
def ticket_report() -> str:
    start = request.args.get('start', (date.today() - timedelta(days=30)).isoformat())
    end = request.args.get('end', date.today().isoformat())
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end) + timedelta(days=1)

    query = Ticket.query.filter(Ticket.created_at >= start_dt, Ticket.created_at < end_dt)

    total = query.count()
    resolved = query.filter(Ticket.status.in_(['resolved', 'closed'])).count()
    breached = query.filter_by(sla_breached=True).count()

    by_priority = dict(db.session.query(Ticket.priority, func.count(Ticket.id))
                       .filter(Ticket.created_at >= start_dt, Ticket.created_at < end_dt)
                       .group_by(Ticket.priority).all())

    by_status = dict(db.session.query(Ticket.status, func.count(Ticket.id))
                     .filter(Ticket.created_at >= start_dt, Ticket.created_at < end_dt)
                     .group_by(Ticket.status).all())

    by_category = dict(db.session.query(Ticket.category, func.count(Ticket.id))
                       .filter(Ticket.created_at >= start_dt, Ticket.created_at < end_dt)
                       .group_by(Ticket.category).all())

    avg_rating = db.session.query(func.avg(Ticket.rating)).filter(
        Ticket.created_at >= start_dt, Ticket.created_at < end_dt, Ticket.rating.isnot(None)
    ).scalar()

    tickets = (Ticket.query.filter(Ticket.created_at >= start_dt, Ticket.created_at < end_dt)
               .order_by(Ticket.created_at.desc()).limit(100).all())

    return render_template('reports/tickets.html',
                           start=start, end=end,
                           total=total, resolved=resolved, breached=breached,
                           resolution_rate=round(resolved / total * 100, 1) if total else 0,
                           breach_rate=round(breached / total * 100, 1) if total else 0,
                           by_priority=by_priority,
                           by_status=by_status,
                           by_category=by_category,
                           avg_rating=round(avg_rating, 2) if avg_rating else 'N/A',
                           tickets=tickets)


@reports_bp.route('/engineers')
@login_required
@admin_or_engineer
def engineer_report() -> str:
    start = request.args.get('start', (date.today() - timedelta(days=30)).isoformat())
    end = request.args.get('end', date.today().isoformat())
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end) + timedelta(days=1)

    engineers = User.query.filter_by(role='engineer', is_active=True).all()
    engineer_stats = []
    for eng in engineers:
        total = Ticket.query.filter_by(assigned_to=eng.id).filter(
            Ticket.created_at >= start_dt, Ticket.created_at < end_dt).count()
        resolved = Ticket.query.filter_by(assigned_to=eng.id, status='resolved').filter(
            Ticket.resolved_at >= start_dt, Ticket.resolved_at < end_dt).count()
        breached = Ticket.query.filter_by(assigned_to=eng.id, sla_breached=True).filter(
            Ticket.created_at >= start_dt, Ticket.created_at < end_dt).count()
        avg_r = db.session.query(func.avg(Ticket.rating)).filter_by(assigned_to=eng.id).filter(
            Ticket.rating.isnot(None), Ticket.created_at >= start_dt, Ticket.created_at < end_dt).scalar()
        engineer_stats.append({
            'engineer': eng,
            'total': total,
            'resolved': resolved,
            'breached': breached,
            'avg_rating': round(avg_r, 2) if avg_r else 'N/A'
        })

    return render_template('reports/engineers.html',
                           engineer_stats=engineer_stats, start=start, end=end)


@reports_bp.route('/attendance')
@login_required
@admin_or_engineer
def attendance_report() -> str:
    start = request.args.get('start', (date.today() - timedelta(days=30)).isoformat())
    end = request.args.get('end', date.today().isoformat())
    start_date = date.fromisoformat(start)
    end_date = date.fromisoformat(end)

    engineers = User.query.filter_by(role='engineer', is_active=True).all()
    att_data = []
    for eng in engineers:
        records = Attendance.query.filter(
            Attendance.engineer_id == eng.id,
            Attendance.work_date >= start_date,
            Attendance.work_date <= end_date
        ).order_by(Attendance.work_date).all()
        total_hours = sum(r.working_hours() for r in records)
        present = len([r for r in records if r.status == 'present'])
        att_data.append({'engineer': eng, 'records': records,
                         'total_hours': round(total_hours, 2), 'present_days': present})

    return render_template('reports/attendance.html',
                           att_data=att_data, start=start, end=end)


@reports_bp.route('/export/tickets/excel')
@login_required
@admin_or_engineer
def export_tickets_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    start = request.args.get('start', (date.today() - timedelta(days=30)).isoformat())
    end = request.args.get('end', date.today().isoformat())
    start_dt = datetime.fromisoformat(start)
    end_dt = datetime.fromisoformat(end) + timedelta(days=1)

    tickets = Ticket.query.filter(Ticket.created_at >= start_dt, Ticket.created_at < end_dt).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Tickets'

    headers = ['Ticket #', 'Title', 'Priority', 'Category', 'Status', 'Created By',
               'Assigned To', 'SLA Breached', 'Created At', 'Resolved At', 'Rating']
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, t in enumerate(tickets, 2):
        ws.cell(row=row, column=1, value=t.ticket_number)
        ws.cell(row=row, column=2, value=t.title)
        ws.cell(row=row, column=3, value=t.priority.title())
        ws.cell(row=row, column=4, value=t.category.title())
        ws.cell(row=row, column=5, value=t.status.replace('_', ' ').title())
        ws.cell(row=row, column=6, value=t.creator.full_name if t.creator else '')
        ws.cell(row=row, column=7, value=t.assignee.full_name if t.assignee else 'Unassigned')
        ws.cell(row=row, column=8, value='Yes' if t.sla_breached else 'No')
        ws.cell(row=row, column=9, value=t.created_at.strftime('%Y-%m-%d %H:%M') if t.created_at else '')
        ws.cell(row=row, column=10, value=t.resolved_at.strftime('%Y-%m-%d %H:%M') if t.resolved_at else '')
        ws.cell(row=row, column=11, value=t.rating or '')

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 18

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'tickets_{start}_{end}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@reports_bp.route('/export/attendance/excel')
@login_required
@admin_or_engineer
def export_attendance_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    start = request.args.get('start', (date.today() - timedelta(days=30)).isoformat())
    end = request.args.get('end', date.today().isoformat())
    start_date = date.fromisoformat(start)
    end_date = date.fromisoformat(end)

    records = (Attendance.query
               .join(User, Attendance.engineer_id == User.id)
               .filter(Attendance.work_date >= start_date, Attendance.work_date <= end_date)
               .order_by(User.full_name, Attendance.work_date).all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Attendance'
    headers = ['Engineer', 'Date', 'Check In', 'Check Out', 'Hours', 'Status']
    header_fill = PatternFill("solid", fgColor="1a1a2e")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, r in enumerate(records, 2):
        ws.cell(row=row, column=1, value=r.engineer.full_name)
        ws.cell(row=row, column=2, value=str(r.work_date))
        ws.cell(row=row, column=3, value=r.check_in.strftime('%H:%M') if r.check_in else '')
        ws.cell(row=row, column=4, value=r.check_out.strftime('%H:%M') if r.check_out else '')
        ws.cell(row=row, column=5, value=r.working_hours())
        ws.cell(row=row, column=6, value=r.status.title())

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = 16

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(output, download_name=f'attendance_{start}_{end}.xlsx',
                     as_attachment=True, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
